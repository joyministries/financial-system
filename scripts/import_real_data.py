#!/usr/bin/env python3
"""
Import real school data from Excel class lists and CSV financial statements.

Usage: PYTHONUNBUFFERED=1 python3 scripts/import_real_data.py
"""

import csv
import re
import uuid
import secrets
from datetime import datetime, UTC
from decimal import Decimal
import openpyxl
import psycopg2
import psycopg2.extras

# ── Config ──────────────────────────────────────────
DB_URL = "postgresql://admin:REDACTED_PASSWORD@dpg-d9s91g2jnfac73932et0-a.ohio-postgres.render.com/financial_db_r82w"
EXCEL_PATH = "/home/unknwn/Downloads/2026 CLASS LISTS.xlsx"
CSV_PATH = "/home/unknwn/Downloads/CustomerStatementReport (1).csv"

GRADE_FEES = {
    'GRADE RR': {'monthly': 1700, 'annual': 20400},
    'GRADE R':  {'monthly': 1700, 'annual': 20400},
    'GRADE 1':  {'monthly': 1940, 'annual': 23280},
    'GRADE 2':  {'monthly': 1940, 'annual': 23280},
    'GRADE 3':  {'monthly': 1940, 'annual': 23280},
    'GRADE 4':  {'monthly': 2250, 'annual': 27000},
    'GRADE 5':  {'monthly': 2250, 'annual': 27000},
    'GRADE 6':  {'monthly': 2250, 'annual': 27000},
    'GRADE 7':  {'monthly': 2380, 'annual': 28560},
    'GRADE 8':  {'monthly': 2980, 'annual': 35760},
    'GRADE 9':  {'monthly': 3230, 'annual': 38760},
}

SHEET_TO_GRADE = {
    'Grade RR': 'GRADE RR', 'Grade R': 'GRADE R',
    'Grade 1 Bethlehem': 'GRADE 1', 'Grade 2 Jericho': 'GRADE 2',
    'Grade 3 ENOCH': 'GRADE 3', ' Grade 3 Enoch': 'GRADE 3',
    'Grade 4 Capernaum': 'GRADE 4', 'Grade 3 Jerusalem': 'GRADE 3',
    'Grade 4 Hebron': 'GRADE 4', 'Grade 5 Babel ': 'GRADE 5',
    'Grade 6 Israel': 'GRADE 6', 'Grade 7 Zion': 'GRADE 7',
    'Grade 8 Jordan': 'GRADE 8', 'Grade 9 Bethel': 'GRADE 9',
}


def hash_password(password: str) -> str:
    import bcrypt
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    students = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        target_grade = SHEET_TO_GRADE.get(sheet_name, sheet_name)
        for row in ws.iter_rows(min_row=5, values_only=True):
            num, student_no, status, full_names, grade, main_contact, other_contact, email = row[:8]
            if student_no and full_names:
                students.append({
                    'student_no': str(int(student_no)) if isinstance(student_no, (int, float)) else str(student_no),
                    'name': str(full_names).strip(),
                    'grade': target_grade,
                    'phone': str(main_contact).strip() if main_contact else None,
                    'other_phone': str(other_contact).strip() if other_contact else None,
                    'email': str(email).strip() if email and str(email).strip() != 'None' else None,
                    'status': str(status).strip() if status else 'Active',
                })
    return students


def parse_csv():
    with open(CSV_PATH, 'r', newline='') as f:
        content = f.read()
    statement_blocks = content.split('STATEMENT')
    parsed = []
    for block in statement_blocks[1:]:
        lines = block.strip().split('\n')
        name_match = re.search(r'\((\d+)\)\s+([A-Z,\s]+?)(?:\s*")', block)
        if not name_match:
            name_match = re.search(r'\((\d+)\)\s+(.+?)[""\n]', block)
        if not name_match:
            continue
        student_no = name_match.group(1)
        parent_name_raw = name_match.group(2).strip().rstrip('",')
        if 'VAT NO' in parent_name_raw:
            parent_name_raw = parent_name_raw.split('VAT NO')[0].strip()
        parent_name_raw = parent_name_raw.strip(' ,')
        if ',' in parent_name_raw:
            parts = parent_name_raw.split(',', 1)
            parent_last = parts[0].strip().title()
            parent_first = parts[1].strip().title()
            parent_name = f"{parent_first} {parent_last}"
        else:
            parent_name = parent_name_raw.title()
        amount_due = None
        amount_paid = None
        for line in lines:
            if 'Amount Due for 2026' in line:
                m = re.search(r'R\s+([\d,]+\.?\d*)', line)
                if m:
                    amount_due = float(m.group(1).replace(',', ''))
            if 'Amount Paid to date' in line:
                m = re.search(r'R\s+([\d,]+\.?\d*)', line)
                if m:
                    amount_paid = float(m.group(1).replace(',', ''))
        # Use csv.reader for proper quoted field handling
        transactions = []
        # Find lines that start with a date pattern (transaction lines)
        for line in lines:
            stripped = line.strip()
            if not re.match(r'"?(\d{2}/\d{2}/\d{4})', stripped):
                continue
            # Strip outer quotes and unescape doubled quotes for csv.reader
            if stripped.startswith('"') and stripped.endswith('"'):
                stripped = stripped[1:-1].replace('""', '"')
            try:
                row = next(csv.reader([stripped]))
            except Exception:
                continue
            date_str = row[0].strip()
            if not re.match(r'\d{2}/\d{2}/\d{4}', date_str):
                continue
            ref = row[2].strip() if len(row) > 2 else ''
            desc = row[3].strip() if len(row) > 3 else ''
            # Columns: 0=Date, 1=empty, 2=Reference, 3=Description, 4=empty, 5=Debit, 6=Credit
            debit = None
            credit = None
            debit_str = row[5].strip() if len(row) > 5 else ''
            credit_str = row[6].strip() if len(row) > 6 else ''
            m_debit = re.match(r'^R\s+([\d,]+\.?\d*)$', debit_str)
            m_credit = re.match(r'^R\s+([\d,]+\.?\d*)$', credit_str)
            if m_debit:
                debit = float(m_debit.group(1).replace(',', ''))
            if m_credit:
                credit = float(m_credit.group(1).replace(',', ''))
            transactions.append({'date': date_str, 'ref': ref, 'description': desc, 'debit': debit, 'credit': credit})
        parsed.append({'student_no': student_no, 'parent_name': parent_name, 'transactions': transactions, 'amount_due': amount_due, 'amount_paid': amount_paid})
    return parsed


def main():
    print("=" * 60)
    print("LCS Financial System - Real Data Import")
    print("=" * 60)

    print("\n[1/7] Parsing Excel class lists...")
    students = parse_excel()
    print(f"  Found {len(students)} students")

    print("\n[2/7] Parsing CSV financial statements...")
    statements = parse_csv()
    print(f"  Found {len(statements)} student statements")

    print("\n[3/7] Connecting to database...")
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    cur = conn.cursor()

    cur.execute("SELECT id FROM users WHERE role IN ('admin', 'super_admin') LIMIT 1")
    admin_id = cur.fetchone()[0]

    try:
        # ── Fee structures ──────────────────────────
        print("\n[4/7] Updating fee structures...")
        cur.execute("SELECT id, name FROM grades")
        grade_map = {row[1]: row[0] for row in cur.fetchall()}

        for grade_name, fees in GRADE_FEES.items():
            grade_id = grade_map.get(grade_name)
            if not grade_id:
                print(f"  WARNING: Grade '{grade_name}' not found in DB")
                continue
            cur.execute("SELECT id FROM fee_structures WHERE grade_id = %s AND academic_year = 2026 AND category = 'Tuition'", (grade_id,))
            existing = cur.fetchone()
            if existing:
                cur.execute("UPDATE fee_structures SET annual_amount = %s, monthly_installment = %s, updated_at = NOW() WHERE id = %s",
                            (fees['annual'], fees['monthly'], existing[0]))
                print(f"  Updated {grade_name}: R{fees['monthly']}/mo")
            else:
                cur.execute("INSERT INTO fee_structures (id, grade_id, academic_year, category, annual_amount, monthly_installment, is_active, payment_plan, created_at, updated_at) VALUES (%s, %s, 2026, 'Tuition', %s, %s, true, 'monthly', NOW(), NOW())",
                            (str(uuid.uuid4()), grade_id, fees['annual'], fees['monthly']))
                print(f"  Created {grade_name}: R{fees['monthly']}/mo")
        conn.commit()

        # ── Parents and students ────────────────────
        print("\n[5/7] Creating parent accounts and students...")
        parent_groups = {}
        for s in students:
            key = s['email'] or s['phone'] or f"parent_{s['student_no']}"
            parent_groups.setdefault(key, []).append(s)
        print(f"  {len(parent_groups)} parent groups identified")

        cur.execute("SELECT email FROM users")
        existing_emails = set(row[0] for row in cur.fetchall())

        cur.execute("SELECT student_number FROM students")
        existing_student_nos = set(row[0] for row in cur.fetchall())

        parent_count = 0
        student_count = 0
        temp_passwords = []
        user_rows = []
        student_rows = []
        guardian_rows = []

        for key, group_students in parent_groups.items():
            first_student = group_students[0]
            parent_email = first_student['email'] or f"parent_{first_student['student_no']}@lcs.school"

            if parent_email in existing_emails:
                cur.execute("SELECT id FROM users WHERE email = %s", (parent_email,))
                parent_id = cur.fetchone()[0]
            else:
                parent_id = str(uuid.uuid4())
                temp_password = f"Temp{secrets.token_hex(3)}!"

                # Use parent name from CSV if available
                csv_stmt = next((s for s in statements if s['student_no'] == first_student['student_no']), None)
                if csv_stmt and csv_stmt['parent_name']:
                    parent_name = csv_stmt['parent_name']
                else:
                    name_parts = first_student['name'].split()
                    parent_name = f"{name_parts[0]} {' '.join(name_parts[1:])}" if len(name_parts) >= 2 else first_student['name']

                user_rows.append((parent_id, parent_email, parent_name.strip(), 'parent', first_student['phone'], True, hash_password(temp_password)))
                existing_emails.add(parent_email)

                temp_passwords.append({'email': parent_email, 'password': temp_password, 'name': parent_name.strip(), 'students': [s['name'] for s in group_students]})
                parent_count += 1

            for s in group_students:
                if s['student_no'] in existing_student_nos:
                    continue
                grade_id = grade_map.get(s['grade'])
                if not grade_id:
                    continue
                student_id = str(uuid.uuid4())
                name_parts = s['name'].split()
                first_name = name_parts[0] if name_parts else s['name']
                last_name = ' '.join(name_parts[1:]) if len(name_parts) >= 2 else ''
                student_rows.append((student_id, s['student_no'], first_name, last_name, grade_id, parent_id))
                guardian_rows.append((str(uuid.uuid4()), student_id, 'primary', f"{first_name} {last_name}".strip(), s['phone'], s['email']))
                existing_student_nos.add(s['student_no'])
                student_count += 1

        if user_rows:
            psycopg2.extras.execute_batch(cur, """
                INSERT INTO users (id, email, full_name, role, phone, is_active, hashed_password, created_at, updated_at)
                VALUES (%s, %s, %s, 'parent', %s, %s, %s, NOW(), NOW())
            """, user_rows, page_size=500)
        if student_rows:
            psycopg2.extras.execute_batch(cur, """
                INSERT INTO students (id, student_number, first_name, last_name, grade_id, parent_id, enrollment_date, is_active, registration_status, payment_preference, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, '2026-01-01', true, 'approved', 'monthly', NOW(), NOW())
            """, student_rows, page_size=500)
        if guardian_rows:
            psycopg2.extras.execute_batch(cur, """
                INSERT INTO student_guardians (id, student_id, guardian_type, full_name, phone, email, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, NOW())
            """, guardian_rows, page_size=500)

        conn.commit()
        print(f"  Created {parent_count} parent accounts")
        print(f"  Created {student_count} student records")

        # ── Financial data ──────────────────────────
        print("\n[6/7] Importing financial data...")

        # Pre-load maps
        cur.execute("SELECT id, student_number, grade_id FROM students")
        student_map = {row[1]: (row[0], row[2]) for row in cur.fetchall()}

        cur.execute("SELECT id, grade_id, monthly_installment FROM fee_structures WHERE academic_year = 2026 AND category = 'Tuition'")
        fee_map = {row[1]: (row[0], row[2]) for row in cur.fetchall()}

        cur.execute("SELECT student_id, EXTRACT(MONTH FROM payment_date)::int FROM payments WHERE EXTRACT(YEAR FROM payment_date) = 2026")
        payment_months = set((row[0], row[1]) for row in cur.fetchall())

        cur.execute("SELECT id, fee_structure_id, month FROM monthly_schedules WHERE academic_year = 2026")
        schedule_map = {(row[1], row[2]): row[0] for row in cur.fetchall()}

        cur.execute("SELECT invoice_number FROM invoices")
        existing_invoices = set(row[0] for row in cur.fetchall())

        payment_rows = []
        receipt_rows = []
        invoice_rows = []
        schedule_rows = []
        outstanding_rows = []
        payment_count = 0
        invoice_count = 0
        outstanding_count = 0

        for stmt in statements:
            sinfo = student_map.get(stmt['student_no'])
            if not sinfo:
                continue
            student_id, grade_id = sinfo
            fin = fee_map.get(grade_id)
            monthly_amount = fin[1] if fin else 0
            fee_struct_id = fin[0] if fin else None

            for txn in stmt['transactions']:
                if txn['credit'] and txn['credit'] > 0:
                    payment_id = str(uuid.uuid4())
                    try:
                        payment_date = datetime.strptime(txn['date'], '%d/%m/%Y')
                    except Exception:
                        continue
                    now = datetime.now(UTC)
                    receipt_number = f"RCP-{stmt['student_no']}-{txn['ref'][-7:]}" if txn['ref'] else f"RCP-{payment_id[:8]}"
                    payment_rows.append((payment_id, student_id, txn['credit'], 'bank_transfer', payment_date, txn['ref'] or '', 'verified', admin_id, now, now))
                    receipt_rows.append((str(uuid.uuid4()), receipt_number, payment_id, student_id, txn['credit'], 'bank_transfer', admin_id, now))
                    payment_months.add((student_id, payment_date.month))
                    payment_count += 1

                elif txn['debit'] and txn['debit'] > 0:
                    try:
                        inv_date = datetime.strptime(txn['date'], '%d/%m/%Y')
                    except Exception:
                        continue
                    invoice_number = f"INV-{stmt['student_no']}-{inv_date.strftime('%Y%m')}"
                    if invoice_number in existing_invoices:
                        continue
                    existing_invoices.add(invoice_number)
                    now = datetime.now(UTC)
                    invoice_rows.append((str(uuid.uuid4()), invoice_number, student_id, 2026, inv_date.month, inv_date, inv_date.replace(day=28), txn['debit'], 0, txn['debit'], 'issued', admin_id, now, now))
                    invoice_count += 1

            # Outstanding balances
            if stmt['amount_due'] and stmt['amount_due'] > 0 and fee_struct_id:
                now = datetime.now(UTC)
                for month in range(1, 9):
                    if (student_id, month) in payment_months:
                        continue
                    sched_key = (fee_struct_id, month)
                    schedule_id = schedule_map.get(sched_key)
                    if not schedule_id:
                        schedule_id = str(uuid.uuid4())
                        schedule_rows.append((schedule_id, fee_struct_id, month, 2026, monthly_amount, datetime(2026, month, 28, tzinfo=UTC), False, now))
                        schedule_map[sched_key] = schedule_id
                    outstanding_rows.append((str(uuid.uuid4()), student_id, schedule_id, monthly_amount, 0, 0, monthly_amount, 'pending', now, now))
                    outstanding_count += 1

            # Flush every 500 payments
            if len(payment_rows) >= 500:
                _flush(conn, cur, payment_rows, receipt_rows, invoice_rows, schedule_rows, outstanding_rows)
                payment_rows, receipt_rows, invoice_rows, schedule_rows, outstanding_rows = [], [], [], [], []
                print(f"  ... flushed (payments: {payment_count}, invoices: {invoice_count})")

        _flush(conn, cur, payment_rows, receipt_rows, invoice_rows, schedule_rows, outstanding_rows)
        conn.commit()
        print(f"  Created {payment_count} payment records")
        print(f"  Created {invoice_count} invoice records")
        print(f"  Created {outstanding_count} outstanding balances")

        # ── Summary ─────────────────────────────────
        print("\n[7/7] Import complete!")
        print(f"\n{'=' * 60}")
        print(f"SUMMARY")
        print(f"{'=' * 60}")
        print(f"Students imported: {student_count}")
        print(f"Parent accounts: {parent_count}")
        print(f"Payments imported: {payment_count}")
        print(f"Invoices created: {invoice_count}")
        print(f"Outstanding balances: {outstanding_count}")

        with open('/home/unknwn/finanacial-system/scripts/temp_credentials.txt', 'w') as f:
            f.write("LCS Parent Account Credentials\n")
            f.write("=" * 60 + "\n")
            f.write("These are temporary passwords. Parents must change them on first login.\n\n")
            for cred in sorted(temp_passwords, key=lambda x: x['email']):
                f.write(f"Email: {cred['email']}\n")
                f.write(f"Password: {cred['password']}\n")
                f.write(f"Parent Name: {cred['name']}\n")
                f.write(f"Children: {', '.join(cred['students'])}\n")
                f.write("-" * 40 + "\n")
        print(f"\nTemporary credentials saved to: scripts/temp_credentials.txt")
        print(f"Total parent accounts with temp passwords: {len(temp_passwords)}")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        cur.close()
        conn.close()


def _flush(conn, cur, payments, receipts, invoices, schedules, outstanding):
    """Batch-insert all accumulated rows."""
    if payments:
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO payments (id, student_id, amount, payment_method, payment_date, reference_number, status, allocated_by, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, payments, page_size=500)
    if receipts:
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO receipts (id, receipt_number, payment_id, student_id, amount, payment_method, allocated_by, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, receipts, page_size=500)
    if invoices:
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO invoices (id, invoice_number, student_id, academic_year, month, issue_date, due_date, subtotal, amount_paid, balance_due, status, created_by, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, invoices, page_size=500)
    if schedules:
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO monthly_schedules (id, fee_structure_id, month, academic_year, amount_due, due_date, is_paid, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, schedules, page_size=500)
    if outstanding:
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO outstanding_balances (id, student_id, monthly_schedule_id, original_amount, rollover_amount, amount_paid, balance, status, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, outstanding, page_size=500)
    conn.commit()


if __name__ == '__main__':
    main()
